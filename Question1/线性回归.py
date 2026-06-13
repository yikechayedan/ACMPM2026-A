"""
data_for_question1.py
问题1：多元线性回归预测出厂水浊度 NTU (17变量)
数据：train_for_question1.csv + 自构造9个时滞变量
训练：2025全年，验证：2026年1月，预测：2026年2月1/10/20日
"""
import pandas as pd, numpy as np, os
from math import erf

def t_pvalue(t):
    return 2*(1-0.5*(1+erf(abs(t)/np.sqrt(2))))

def ols(X, y, names):
    n, p = X.shape
    Xd = np.column_stack([np.ones(n), X])
    b = np.linalg.inv(Xd.T @ Xd) @ Xd.T @ y
    yp = Xd @ b; res = y - yp
    ssr, sst = np.sum(res**2), np.sum((y-y.mean())**2)
    r2 = 1-ssr/sst; adj_r2 = 1-(1-r2)*(n-1)/(n-p-1)
    rmse = np.sqrt(ssr/n); mae = np.mean(np.abs(res))
    sigma2 = ssr/(n-p-1); C = np.linalg.inv(Xd.T @ Xd)
    se = np.sqrt(sigma2*np.diag(C)); t = b/se
    pv = np.array([t_pvalue(tv) for tv in t])
    f = (sst-ssr)/p/sigma2; dw = np.sum(np.diff(res)**2)/np.sum(res**2)
    return {'n':n,'p':p,'r2':r2,'adj_r2':adj_r2,'rmse':rmse,'mae':mae,
            'f':f,'dw':dw,'b':b,'t':t,'pv':pv,'yp':yp,'names':['(截距)']+list(names)}

def print_ols(r, title=''):
    print(f'\n--- {title} ---')
    print(f'  n={r["n"]} p={r["p"]}  R2={r["r2"]:.4f}  adjR2={r["adj_r2"]:.4f}  RMSE={r["rmse"]:.4f}  MAE={r["mae"]:.4f}  DW={r["dw"]:.4f}')
    print(f'  {"变量":<22s} {"系数":>10s} {"t值":>8s} {"p值":>10s}')
    for i in range(len(r['names'])):
        sig = '***' if r['pv'][i]<.001 else ('**' if r['pv'][i]<.01 else ('*' if r['pv'][i]<.05 else ''))
        print(f'  {r["names"][i]:<22s} {r["b"][i]:10.4f} {r["t"][i]:8.3f} {r["pv"][i]:10.4f}  {sig}')

BASE = r'D:\abetwithyjz\A题 自来水厂水质预测与评估'

# ============================================================
# 1. 加载训练数据 + 构造9个时滞变量
# ============================================================
df = pd.read_csv(os.path.join(BASE, '附件/train_for_question1.csv'))
target = 'NTU'
curr_vars = ['RIVER LEVEL','R/W FLOW','R/W NTU','R/W CLR','R/W PH',
             'FILT. NTU','C/W WELL LEVEL','ALUM']

all_vars = list(curr_vars)
for col in ['R/W NTU','FILT. NTU','ALUM']:
    for lag in [1,2,3]:
        all_vars.append(f'{col}_lag{lag}')
        df[f'{col}_lag{lag}'] = df[col].shift(lag)

df = df.iloc[3:].copy()  # 剔除lag产生的NaN
X = df[all_vars].values
y = df[target].values
print(f'训练集: {len(df)}行, {len(all_vars)}变量, NTU mean={y.mean():.4f} std={y.std():.4f}')

# ============================================================
# 2. OLS 全变量回归 + 特征重要性
# ============================================================
print('\n' + '='*60)
print('OLS 多元线性回归 (17变量)')
print('='*60)
r = ols(X, y, all_vars)
print_ols(r)

# 标准化系数排序
std_coef = sorted([(all_vars[j], abs(r['b'][j+1])*X[:,j].std()) for j in range(len(all_vars))],
                  key=lambda x: x[1], reverse=True)
print(f'\n特征重要性 (|beta*std|, 前10):')
for rank, (name, imp) in enumerate(std_coef[:10], 1):
    print(f'  {rank:2d}. {name:<24s} {imp:10.4f}')

# ============================================================
# 3. 验证: 2026年1月
# ============================================================
print('\n' + '='*60)
print('验证: 2026年1月')
print('='*60)

def load_month(month):
    path = os.path.join(BASE, f'附件/附件2  2026数据集/2026年{month}月.xlsx')
    xl = pd.ExcelFile(path)
    sheets = sorted(xl.sheet_names, key=lambda s: tuple(int(x.strip()) for x in s.strip().split('.')))
    all_days = []
    for sh in sheets:
        d = pd.read_excel(path, sheet_name=sh)
        d.columns = [c.strip() for c in d.columns]
        d.insert(0, 'DATE', f'2026-{month:02d}-{int(sh.strip().split(".")[0]):02d}')
        all_days.append(d)
    return pd.concat(all_days, ignore_index=True)

def prep_month(df_m, tail_df):
    for col in curr_vars:
        if col in df_m.columns:
            df_m[col] = pd.to_numeric(df_m[col].replace('-',np.nan), errors='coerce')
        df_m[col] = df_m[col].ffill().bfill()
    for col in ['R/W NTU','FILT. NTU','ALUM']:
        s_all = pd.concat([tail_df[col], df_m[col].reset_index(drop=True)], ignore_index=True)
        for lag in [1,2,3]:
            df_m[f'{col}_lag{lag}'] = s_all.shift(lag).iloc[3:].values
    for col in all_vars:
        df_m[col] = df_m[col].ffill().bfill()
    return df_m

tail = df[['R/W NTU','FILT. NTU','ALUM']].iloc[-3:].reset_index(drop=True)

df_jan = prep_month(load_month(1), tail)
df_jv = df_jan.loc[df_jan[target].notnull()].iloc[3:].copy()
Xv, yv = df_jv[all_vars].values, df_jv[target].values
yp_v = np.column_stack([np.ones(len(Xv)), Xv]) @ r['b']
ssr_v, sst_v = np.sum((yv-yp_v)**2), np.sum((yv-yv.mean())**2)
print(f'验证集: {len(yv)}行  MAE={np.mean(np.abs(yv-yp_v)):.4f}  RMSE={np.sqrt(ssr_v/len(yv)):.4f}  R2={1-ssr_v/sst_v:.4f}')
print(f'训练NTU std={y.std():.4f}  验证NTU std={yv.std():.4f}')

# ============================================================
# 4. 预测: 2026年2月1/10/20日
# ============================================================
print('\n' + '='*60)
print('预测: 2026年2月')
print('='*60)

feb_path = os.path.join(BASE, '附件/附件2  2026数据集/2026年2月.xlsx')
xl_feb = pd.ExcelFile(feb_path)
target_sheets = {'02-01': '01.02', '02-10': '10.02', '02-20': '20.02'}
results = []

for label, sh in target_sheets.items():
    df_day = pd.read_excel(feb_path, sheet_name=sh)
    df_day.columns = [c.strip() for c in df_day.columns]
    time_col = [c for c in df_day.columns if 'TIME' in c.upper()][0]
    df_day = prep_month(df_day, tail)
    df_day = df_day.sort_values(time_col).reset_index(drop=True)
    pred = np.maximum(np.column_stack([np.ones(len(df_day)), df_day[all_vars].values]) @ r['b'], 0)
    results.append((sh.strip(), dict(zip(df_day[time_col], pred))))

    print(f'\n{label} 日均={pred.mean():.4f} 最小={pred.min():.4f} 最大={pred.max():.4f}')
    for i in range(12):
        print(f'  {int(df_day[time_col].iloc[i]):04d}  {pred[i]:.4f}')

# ============================================================
# 5. 保存
# ============================================================
print('\n>>> 保存预测结果...')
pred_dict = {s: p for s, p in results}
out_path = os.path.join(BASE, '附件', '问题1_NTU预测结果.xlsx')

from openpyxl import Workbook
wb = Workbook(); wb.remove(wb.active)
for sh in xl_feb.sheet_names:
    df_orig = pd.read_excel(feb_path, sheet_name=sh)
    sh_clean = sh.strip()
    time_col = [c for c in df_orig.columns if 'TIME' in c.upper()][0]
    if sh_clean in pred_dict:
        df_orig['NTU'] = df_orig[time_col].map(pred_dict[sh_clean])
        df_orig = df_orig.sort_values(time_col)
    ws = wb.create_sheet(title=sh_clean)
    for c, cn in enumerate(df_orig.columns, 1):
        ws.cell(row=1, column=c, value=cn)
    for ri, row in enumerate(df_orig.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=ri, column=c, value=val if pd.notna(val) else None)
wb.save(out_path)
print(f'已保存: {out_path}')

print('\n' + '='*60)
print('汇总')
print('='*60)
print(f'{"":<12s} {"MAE":>8s} {"RMSE":>8s} {"R2":>8s}')
print(f'{"训练":<12s} {r["mae"]:8.4f} {r["rmse"]:8.4f} {r["r2"]:8.4f}')
print(f'{"验证":<12s} {np.mean(np.abs(yv-yp_v)):8.4f} {np.sqrt(ssr_v/len(yv)):8.4f} {1-ssr_v/sst_v:8.4f}')
print('\n完成。')
